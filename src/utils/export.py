"""Export utilities for data export to different formats."""
import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class DataExporter:
    """Data export manager."""
    
    def __init__(self, export_dir: str = "exports"):
        """Initialize exporter.
        
        Args:
            export_dir: Directory for exported files
        """
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
    def export_to_csv(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Export data to CSV format.
        
        Args:
            data: List of dictionaries containing data
            filename: Name of the output file
            
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
            
        output_file = self.export_dir / f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
            
        return output_file
        
    def export_to_excel(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Export data to Excel format.
        
        Args:
            data: List of dictionaries containing data
            filename: Name of the output file
            
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
            
        output_file = self.export_dir / f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Write data
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = item.get(key)
                cell.alignment = Alignment(horizontal="center")
                
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(output_file)
        return output_file
        
    def export_to_json(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Export data to JSON format.
        
        Args:
            data: List of dictionaries containing data
            filename: Name of the output file
            
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
            
        output_file = self.export_dir / f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
            
        return output_file
        
    def export_to_pdf(self, data: List[Dict[str, Any]], filename: str) -> Path:
        """Export data to PDF format using pandas.
        
        Args:
            data: List of dictionaries containing data
            filename: Name of the output file
            
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
            
        output_file = self.export_dir / f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        df = pd.DataFrame(data)
        df.to_html(output_file, index=False)
        
        return output_file
        
    def clear_old_exports(self, days: int = 30) -> None:
        """Clear exported files older than specified days.
        
        Args:
            days: Number of days to keep exports
        """
        cutoff = datetime.now().timestamp() - (days * 24 * 60 * 60)
        
        for export_file in self.export_dir.glob("*"):
            if export_file.stat().st_mtime < cutoff:
                export_file.unlink() 